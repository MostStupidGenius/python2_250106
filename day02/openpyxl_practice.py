# openpyxl_practice.py
# 엑셀을 다루기 위한 라이브러리 -> openpyxl
# pip install openpyxl pandas matplotlib
import openpyxl
import openpyxl.workbook

# 엑셀 파일 생성
# 엑셀의 기본 구조
# 1. 워크북 workbook
# 2. 시트   sheet
# 3. 셀     cell

# 워크북 생성
wb = openpyxl.Workbook()

# 워크북의 활성화된 시트를 가져온다.
sheet = wb.active

# 딕셔너리처럼 문자열로 셀 주소를 전달하여 접근,
# 넣고자 하는 값을 대입
sheet['A1'] = "이름"
sheet['B1'] = "나이"
sheet['A2'] = "홍길동"
sheet['B2'] = 30

file_path ="example.xlsx"

# 시트의 제목을 바꾸는 방법
sheet.title = "info"


# 엑셀 파일 저장
wb.save(file_path)

# 엑셀 파일 읽기
wb = openpyxl.load_workbook(file_path)

sheet = wb.active

# print(sheet['A2'].value) # 출력: 홍길동
# print(sheet['B2'].value) # 출력: 30

# cell의 행번호와 열번호를 이용해서
# 해당 cell의 좌표에 접근할 수도 있다.
print(sheet.cell(2, 1).value)
print(sheet.cell(2, 2).value)







